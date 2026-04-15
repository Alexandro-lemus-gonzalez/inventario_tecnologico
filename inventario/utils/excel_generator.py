import openpyxl
import os
from datetime import datetime
from django.conf import settings
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def exportar_activos_excel(activos):
    """
    Exporta el listado de activos a un archivo Excel con estilos profesionales.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Activos"
    
    # Encabezados
    headers = ['CÓDIGO', 'NOMBRE DEL ACTIVO', 'CATEGORÍA', 'UBICACIÓN', 'ESTADO', 'MARCA', 'MODELO', 'SERIE', 'VALOR ADQ.', 'RESPONSABLE']
    ws.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color='39A900', end_color='39A900', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for a in activos:
        ws.append([
            a.codigo_inventario,
            a.nombre,
            a.categoria.nombre,
            a.ubicacion.nombre if a.ubicacion else 'N/A',
            a.estado.nombre,
            a.marca or 'N/A',
            a.modelo or 'N/A',
            a.numero_serie or 'N/A',
            float(a.valor_adquisicion) if a.valor_adquisicion else 0,
            a.responsable or 'No asignado'
        ])
    
    # Formato moneda para la columna de valor (columna I = 9)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=9).number_format = '"$"#,##0.00'
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 4
        
    filename = f"inventario_activos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def exportar_logs_excel(logs):
    """
    Exporta el historial de auditoría a Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Auditoría"
    
    headers = ['FECHA/HORA', 'USUARIO', 'ACCIÓN', 'MODELO', 'OBJETO', 'DESCRIPCIÓN', 'IP']
    ws.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color='00324D', end_color='00324D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for log in logs:
        ws.append([
            log.fecha_hora.strftime('%d/%m/%Y %H:%M:%S'),
            log.usuario_nombre,
            log.get_accion_display(),
            log.modelo,
            log.objeto_repr,
            log.descripcion,
            log.ip_address or 'N/A'
        ])
        
    filename = f"auditoria_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath
